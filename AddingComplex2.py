import os
import sys
import psutil
import streamlit as st
import pandas as pd
import openpyxl
import glob
import base64
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import time

LOCK_FILE = "streamlit_running.lock"

def start_streamlit():
    script_path = os.path.abspath(__file__)
    command = f'{sys.executable} -m streamlit run "{script_path}"'
    os.system(f'start cmd /k "{command}"')
    print("Streamlit server started in a new command window.")

def run_streamlit_app():
    print("Running Streamlit app...")

    # Boeing color scheme
    BOEING_BLUE = "#0033A0"
    BOEING_LIGHT_BLUE = "#A0CCEB"
    BOEING_GRAY = "#444444"

    # Function to load and display the Boeing logo
    def get_base64_of_bin_file(bin_file):
        with open(bin_file, 'rb') as f:
            data = f.read()
        return base64.b64encode(data).decode()

    def set_boeing_theme():
        # Determine the current directory for the logo
        current_dir = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__)
        logo_path = os.path.join(current_dir, 'Boeing_full_logo.png')
        if not os.path.exists(logo_path):
            st.error("Boeing logo not found in the current directory.")
            return

        logo_base64 = get_base64_of_bin_file(logo_path)

        st.markdown(
            f"""
            <style>
                .reportview-container .main .block-container{{
                    max-width: 1000px;
                    padding-top: 2rem;
                    padding-bottom: 2rem;
                    padding-left: 5rem;
                    padding-right: 5rem;
                }}
                .reportview-container .main {{
                    color: {BOEING_GRAY};
                    background-color: white;
                }}
                .sidebar .sidebar-content {{
                    background-color: {BOEING_LIGHT_BLUE};
                }}
                h1 {{
                    color: {BOEING_BLUE};
                }}
                .stButton>button {{
                    color: white;
                    background-color: {BOEING_BLUE};
                    border-radius: 5px;
                }}
                .stTextInput>div>div>input {{
                    border-color: {BOEING_BLUE};
                }}
            </style>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            f"""
            <img src="data:image/png;base64,{logo_base64}" alt="Boeing Logo" style="width:200px;margin-bottom:20px;">
            """,
            unsafe_allow_html=True,
        )

    # Set Boeing theme
    set_boeing_theme()

    st.title("RCR Sheet Automator")

    # Create input fields
    input1 = st.text_input("Enter the directory path where the files are stored: ")
    input2 = st.text_input("Enter the name of the permanent Excel file: ")
    input3 = st.text_input("Enter the pattern for the monthly files: ")
    input4 = st.text_input("Enter the name of the sheet in the permanent file where the data should be added: ")
    input5 = st.text_input("Enter the starting row number: ")

    # Add a button to submit the inputs
    if st.button("Submit"):
        if input1 and input2 and input3 and input4 and input5:
            try:
                # Set up variables
                directory_path = input1
                permanent_file_name = input2
                monthly_file_pattern = input3
                sheet_name = input4
                start_row = int(input5)
                header_row = 3

                # Create a progress bar
                progress_bar = st.progress(0)
                status_text = st.empty()

                # Update progress
                def update_progress(progress, status):
                    progress_bar.progress(progress)
                    status_text.text(status)

                # Construct the full path for the permanent file
                permanent_file = os.path.join(directory_path, permanent_file_name)

                # Construct a file pattern to match the monthly files
                file_pattern = os.path.join(directory_path, monthly_file_pattern)

                update_progress(0.1, "Looking for monthly files...")

                # Find all files matching the pattern
                monthly_files = glob.glob(file_pattern)

                if not monthly_files:
                    st.error("No monthly files found matching the pattern.")
                    return

                # Find the latest file based on the naming convention
                latest_file = max(monthly_files, key=os.path.getctime)

                update_progress(0.2, f"Reading data from {latest_file}...")

                # Read the monthly data
                try:
                    monthly_data = pd.read_excel(latest_file, engine='openpyxl')
                except Exception as e:
                    st.error(f"Error reading the monthly file: {e}")
                    return

                update_progress(0.3, "Opening permanent file...")

                # Open the permanent file and load the workbook
                try:
                    permanent_wb = openpyxl.load_workbook(permanent_file, keep_vba=True, read_only=False, data_only=False)
                except Exception as e:
                    st.error(f"Error opening the permanent file: {e}")
                    return

                if sheet_name not in permanent_wb.sheetnames:
                    st.error(f"Sheet '{sheet_name}' does not exist in the permanent file.")
                    return

                sheet = permanent_wb[sheet_name]

                update_progress(0.4, "Processing headers...")

                # Get column headers from the permanent file (row 3)
                permanent_headers = [cell.value for cell in sheet[header_row] if cell.value is not None]
                # Get column headers from the monthly file
                monthly_headers = [header.strip().lower() for header in monthly_data.columns]

                update_progress(0.5, "Preparing data for update...")

                # Clear the cells in the specified rows, but keep formatting and equations
                for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        if cell.data_type != 'f':  # Only clear cell value if it is not a formula
                            cell.value = None

                # Prepare data for update
                for col_idx, header in enumerate(permanent_headers):
                    normalized_header = header.strip().lower()
                    if normalized_header in monthly_headers:
                        monthly_col_idx = monthly_headers.index(normalized_header)
                        for row_idx, value in enumerate(monthly_data.iloc[:, monthly_col_idx]):
                            cell = sheet.cell(row=row_idx + start_row, column=col_idx + 1)
                            cell.value = value
                        if normalized_header == "on rcr events report":
                            break

                update_progress(0.6, "Updating data...")

                print(f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{sheet_name}' starting at row {start_row}")

                # Load the Excel file using pandas
                excel_file = pd.ExcelFile(permanent_file)

                # Read the Candidate_List tab
                candidate_df = excel_file.parse('Top 30', header=2)

                # Read the Facilitator List tab
                facilitator_df = excel_file.parse('Facilitator List With Location')

                # Create a new column in the Candidate_List tab to store the closest facilitator name
                candidate_df['Closest Facilitator'] = ''
                geolocator = Nominatim(user_agent='my_app')

                def geocode_with_retry(location):
                    max_retries = 5
                    retries = 0
                    while retries < max_retries:
                        try:
                            return geolocator.geocode(location)
                        except GeocoderTimedOut:
                            retries += 1
                            time.sleep(1)  # Wait for 1 second before retrying
                    return None

                facilitator_locations = {}
                facilitator_count = {}
                facilitator_df = facilitator_df.dropna()
                for facilitator_index, facilitator_row in facilitator_df.iterrows():
                    facilitator_location = facilitator_row['Location']
                    facilitator_name = facilitator_row['Name']
                    facilitator_state, facilitator_city = facilitator_location.split('-')
                    facilitator_location = f"{facilitator_city.strip()}, {facilitator_state.strip()}"
                    facilitator_coordinates = geocode_with_retry(facilitator_location)
                    facilitator_locations[facilitator_name] = facilitator_coordinates
                    facilitator_count[facilitator_name] = 0

                # Iterate over each RCR in the Candidate_List tab
                for index, candidate_row in candidate_df.head(30).iterrows():
                    candidate_location = f"{candidate_row['City']}, {candidate_row['State']}"
                    candidate_coordinates = geocode_with_retry(candidate_location)
                    if candidate_coordinates is None:
                        continue
                    candidate_latitude = candidate_coordinates.latitude
                    candidate_longitude = candidate_coordinates.longitude
                    closest_facilitator = None
                    min_distance = float('inf')

                    # Iterate over each facilitator in the Facilitator List tab
                    for facilitator_index, facilitator_row in facilitator_df.iterrows():
                        facilitator_coordinates = facilitator_locations[facilitator_row['Name']]
                        if facilitator_coordinates is None:
                            continue
                        facilitator_latitude = facilitator_coordinates.latitude
                        facilitator_longitude = facilitator_coordinates.longitude

                        distance = geodesic((candidate_latitude, candidate_longitude), (facilitator_latitude, facilitator_longitude)).miles

                        # Update the closest facilitator if a closer one is found
                        if distance < min_distance and facilitator_count[facilitator_row['Name']] < facilitator_row['RCR_Limit']:
                            min_distance = distance
                            closest_facilitator = facilitator_row['Name']

                    # Assign the closest facilitator name to the corresponding row in the Candidate_List tab
                    if closest_facilitator is not None:
                        candidate_df.at[index, 'Closest Facilitator'] = closest_facilitator
                        facilitator_count[closest_facilitator] = facilitator_count.get(closest_facilitator, 0) + 1
                    else:
                        # Handle the case where no closest facilitator was found
                        print(f"No facilitator found for candidate at index {index}")

                # Select the sheet you want to update
                ws = permanent_wb['Top 30']

                # Add the new column header
                ws.cell(row=3, column=ws.max_column + 1, value='Closest Facilitator')

                # Write the data to the new column
                for idx, value in enumerate(candidate_df['Closest Facilitator'], start=4):
                    ws.cell(row=idx, column=ws.max_column, value=value)

                # Save the workbook
                permanent_wb.save(permanent_file)

                update_progress(1.0, "Process completed successfully!")
                st.success(f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{sheet_name}' starting at row {start_row}")
                st.success(f"Updated 'Top 30' sheet in '{permanent_file_name}' with closest facilitator information.")

            except Exception as e:
                st.error(f"An error occurred: {str(e)}")
        else:
            st.warning("Please fill in all fields before submitting.")

if __name__ == "__main__":
    if os.path.exists(LOCK_FILE):
        print("Lock file found. Running Streamlit app...")
        try:
            run_streamlit_app()
        finally:
            if os.path.exists(LOCK_FILE):
                os.remove(LOCK_FILE)
    else:
        print("No lock file found. Creating lock file and starting Streamlit...")
        with open(LOCK_FILE, 'w') as f:
            f.write('')
        start_streamlit()
        sys.exit()  # Exit the script after starting Streamlit in a new command window
