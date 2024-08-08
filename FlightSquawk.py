import sys
import os
import pandas as pd
import openpyxl
import glob
import time
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QVBoxLayout, QWidget, QTextEdit, QMessageBox, QProgressBar, QHBoxLayout
from PyQt5.QtGui import QPixmap

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

class RCRAutomatorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('RCR Sheet Automator')
        self.setGeometry(100, 100, 800, 600)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()

        logo_path = resource_path('Boeing_full_logo.png')
        self.logo_label = QLabel(self)
        self.logo_label.setPixmap(QPixmap(logo_path).scaled(200, 100))
        layout.addWidget(self.logo_label)

        # Directory Path Input
        self.input1_label = QLabel("Enter the directory path where the files are stored:")
        layout.addWidget(self.input1_label)
        self.input1 = QLineEdit(self)
        layout.addWidget(self.input1)

        # Permanent File Name Input
        self.input2_label = QLabel("Enter the name of the permanent Excel file:")
        layout.addWidget(self.input2_label)
        self.input2 = QLineEdit(self)
        layout.addWidget(self.input2)

        # Monthly File Pattern Input with Example
        self.input3_label = QLabel("Enter the pattern for the monthly files (e.g., groceries_*.xlsx):")
        layout.addWidget(self.input3_label)
        self.input3 = QLineEdit(self)
        layout.addWidget(self.input3)

        # Sheet Name Input
        self.input4_label = QLabel("Enter the name of the sheet in the permanent file where the data should be added:")
        layout.addWidget(self.input4_label)
        self.input4 = QLineEdit(self)
        layout.addWidget(self.input4)

        # Permanent Header Row Input
        self.input5_label = QLabel("Enter the row number of the headers in the permanent file (e.g., 3):")
        layout.addWidget(self.input5_label)
        self.input5 = QLineEdit(self)
        layout.addWidget(self.input5)

        # Permanent Start Row Input
        self.input6_label = QLabel("Enter the starting row number to populate in the permanent file (e.g., 4):")
        layout.addWidget(self.input6_label)
        self.input6 = QLineEdit(self)
        layout.addWidget(self.input6)

        # Temporary Header Row Input
        self.input7_label = QLabel("Enter the row number of the headers in the temporary file (e.g., 2):")
        layout.addWidget(self.input7_label)
        self.input7 = QLineEdit(self)
        layout.addWidget(self.input7)

        # Temporary Start Row Input
        self.input8_label = QLabel("Enter the starting row number of the data in the temporary file (e.g., 3):")
        layout.addWidget(self.input8_label)
        self.input8 = QLineEdit(self)
        layout.addWidget(self.input8)

        self.submit_button = QPushButton("Submit", self)
        self.submit_button.clicked.connect(self.onSubmit)
        layout.addWidget(self.submit_button)

        self.progress_bar = QProgressBar(self)
        layout.addWidget(self.progress_bar)

        self.status_text = QTextEdit(self)
        self.status_text.setReadOnly(True)
        layout.addWidget(self.status_text)

        central_widget.setLayout(layout)

    def update_progress(self, progress, status):
        self.progress_bar.setValue(progress)
        self.status_text.append(status)

    def onSubmit(self):
        input1 = self.input1.text()
        input2 = self.input2.text()
        input3 = self.input3.text()
        input4 = self.input4.text()
        input5 = self.input5.text()
        input6 = self.input6.text()
        input7 = self.input7.text()
        input8 = self.input8.text()

        if input1 and input2 and input3 and input4 and input5 and input6 and input7 and input8:
            try:
                directory_path = input1
                permanent_file_name = input2
                monthly_file_pattern = input3
                permanent_sheet_name = input4
                permanent_header_row = int(input5)
                permanent_start_row = int(input6)
                temporary_header_row = int(input7)
                temporary_start_row = int(input8)

                self.update_progress(10, "Looking for monthly files...")
                monthly_files = glob.glob(os.path.join(directory_path, monthly_file_pattern))

                if not monthly_files:
                    QMessageBox.critical(self, "Error", "No monthly files found matching the pattern.")
                    return

                latest_file = max(monthly_files, key=os.path.getctime)

                self.update_progress(20, f"Reading data from {latest_file}...")
                try:
                    monthly_data = pd.read_excel(latest_file, header=temporary_header_row - 1, engine='openpyxl')
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Error reading the monthly file: {e}")
                    return

                self.update_progress(30, "Opening permanent file...")
                try:
                    permanent_file = os.path.join(directory_path, permanent_file_name)
                    permanent_wb = openpyxl.load_workbook(permanent_file, keep_vba=True)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Error opening the permanent file: {e}")
                    return

                if permanent_sheet_name not in permanent_wb.sheetnames:
                    QMessageBox.critical(self, "Error", f"Sheet '{permanent_sheet_name}' does not exist in the permanent file.")
                    return

                sheet = permanent_wb[permanent_sheet_name]
                self.update_progress(40, "Processing headers...")

                permanent_headers = [cell.value for cell in sheet[permanent_header_row] if isinstance(cell.value, str)]
                monthly_headers = [header.strip().lower() for header in monthly_data.columns if isinstance(header, str)]

                self.update_progress(50, "Clearing old data...")

                for row in sheet.iter_rows(min_row=permanent_start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        if cell.data_type != 'f':  # Skip clearing if cell contains a formula
                            cell.value = None

                self.update_progress(60, "Updating data...")

                for col_idx, header in enumerate(permanent_headers):
                    normalized_header = header.strip().lower()
                    if normalized_header in monthly_headers:
                        monthly_col_idx = monthly_headers.index(normalized_header)
                        temp_data = monthly_data.iloc[temporary_start_row - temporary_header_row:, monthly_col_idx]
                        for row_offset, value in enumerate(temp_data, start=0):
                            cell = sheet.cell(row=permanent_start_row + row_offset, column=col_idx + 1)
                            cell.value = value

                try:
                    permanent_wb.save(permanent_file)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Error saving the permanent file: {e}")
                    return

                self.update_progress(100, "Process completed successfully!")
                QMessageBox.information(self, "Success", f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{permanent_sheet_name}' starting at row {permanent_start_row}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
        else:
            QMessageBox.warning(self, "Warning", "Please fill in all fields before submitting.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RCRAutomatorApp()
    ex.show()
    sys.exit(app.exec_())
