import pandas as pd
import openpyxl
import os
import glob
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import time

# Get input from the user
directory_path = input("Enter the directory path where the files are stored (e.g., C:\\Users\\YourName\\Documents\\Files): ")
permanent_file_name = input("Enter the name of the permanent Excel file (e.g., permanent_file.xlsm): ")
monthly_file_pattern = input("Enter the pattern for the monthly files (e.g., groceries_*.xlsx): ")
sheet_name = input("Enter the name of the sheet in the permanent file where the data should be added: ")
start_row = int(input("Enter the starting row number (e.g., 4): "))
header_row = 3  # Header row is row 3

# Construct the full path for the permanent file
permanent_file = os.path.join(directory_path, permanent_file_name)

# Construct a file pattern to match the monthly files
file_pattern = os.path.join(directory_path, monthly_file_pattern)

# Find all files matching the pattern
monthly_files = glob.glob(file_pattern)

# Ensure there are files matching the pattern
if not monthly_files:
    print("No monthly files found matching the pattern.")
    exit()

# Find the latest file based on the naming convention
latest_file = max(monthly_files, key=os.path.getctime)

# Verify that the latest file is a valid Excel file
try:
    monthly_data = pd.read_excel(latest_file, engine='openpyxl')
except Exception as e:
    print(f"Error reading the monthly file: {e}")
    exit()

try:
    # Open the permanent file and load the workbook
    permanent_wb = openpyxl.load_workbook(permanent_file, keep_vba=True)
except Exception as e:
    print(f"Error opening the permanent file: {e}")
    exit()

# Access the specific sheet in the permanent workbook
if sheet_name not in permanent_wb.sheetnames:
    print(f"Sheet '{sheet_name}' does not exist in the permanent file.")
    exit()
sheet = permanent_wb[sheet_name]

# Get column headers from the permanent file (row 3)
permanent_headers = [cell.value for cell in sheet[header_row] if cell.value is not None]

# Get column headers from the monthly file
monthly_headers = [header.strip().lower() for header in monthly_data.columns]

# Clear the cells in the specified rows, but keep formatting and equations
for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

# Update the specific sheet starting at the specified row
for col_idx, header in enumerate(permanent_headers):
    normalized_header = header.strip().lower()
    if normalized_header in monthly_headers:
        monthly_col_idx = monthly_headers.index(normalized_header)
        for row_idx, value in enumerate(monthly_data.iloc[:, monthly_col_idx]):
            cell = sheet.cell(row=row_idx + start_row, column=col_idx + 1)
            cell.value = value
        if normalized_header == "on rcr events report":
            break

print(f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{sheet_name}' starting at row {start_row}")

# Load the Excel file using pandas
excel_file = pd.ExcelFile(permanent_file)

# Read the Candidate_List tab
candidate_df = excel_file.parse('Top 30', header=2)

# Read the Facilitator List tab
facilitator_df = excel_file.parse('Facilitator List With Location')

# Create a new column in the Candidate_List tab to store the closest facilitator name
candidate_df['Closest Facilitator'] = ''
geolocator = Nominatim(user_agent='my_app')

def geocode_with_retry(location):
    max_retries = 5
    retries = 0
    while retries < max_retries:
        try:
            return geolocator.geocode(location)
        except GeocoderTimedOut:
            retries += 1
            time.sleep(1)  # Wait for 1 second before retrying
    return None

facilitator_locations = {}
facilitator_count = {}
facilitator_df = facilitator_df.dropna()
for facilitator_index, facilitator_row in facilitator_df.iterrows():
    facilitator_location = facilitator_row['Location']
    facilitator_name = facilitator_row['Name']
    facilitator_state, facilitator_city = facilitator_location.split('-')
    facilitator_location = f"{facilitator_city.strip()}, {facilitator_state.strip()}"
    facilitator_coordinates = geocode_with_retry(facilitator_location)
    facilitator_locations[facilitator_name] = facilitator_coordinates
    facilitator_count[facilitator_name] = 0

# Iterate over each RCR in the Candidate_List tab
for index, candidate_row in candidate_df.head(30).iterrows():
    candidate_location = f"{candidate_row['City']}, {candidate_row['State']}"
    candidate_coordinates = geocode_with_retry(candidate_location)
    if candidate_coordinates is None:
        continue
    candidate_latitude = candidate_coordinates.latitude
    candidate_longitude = candidate_coordinates.longitude
    closest_facilitator = None
    min_distance = float('inf')
    
    # Iterate over each facilitator in the Facilitator List tab
    for facilitator_index, facilitator_row in facilitator_df.iterrows():
        facilitator_coordinates = facilitator_locations[facilitator_row['Name']]
        if facilitator_coordinates is None:
            continue
        facilitator_latitude = facilitator_coordinates.latitude
        facilitator_longitude = facilitator_coordinates.longitude
        
        distance = geodesic((candidate_latitude, candidate_longitude), (facilitator_latitude, facilitator_longitude)).miles
        
        # Update the closest facilitator if a closer one is found
        if distance < min_distance and facilitator_count[facilitator_row['Name']] < facilitator_row['RCR_Limit']:
            min_distance = distance
            closest_facilitator = facilitator_row['Name']
    
    # Assign the closest facilitator name to the corresponding row in the Candidate_List tab
    if closest_facilitator is not None:
        candidate_df.at[index, 'Closest Facilitator'] = closest_facilitator
        facilitator_count[closest_facilitator] = facilitator_count.get(closest_facilitator, 0) + 1
    else:
        # Handle the case where no closest facilitator was found
        print(f"No facilitator found for candidate at index {index}")

# Select the sheet you want to update
ws = permanent_wb['Top 30']

# Add the new column header
ws.cell(row=3, column=ws.max_column + 1, value='Closest Facilitator')

# Write the data to the new column
for idx, value in enumerate(candidate_df['Closest Facilitator'], start=4):
    ws.cell(row=idx, column=ws.max_column, value=value)

# Save the workbook
permanent_wb.save(permanent_file)

print(f"Updated 'Top 30' sheet in '{permanent_file_name}' with closest facilitator information.")
