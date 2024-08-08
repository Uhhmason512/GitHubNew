import streamlit as st
import pandas as pd
import openpyxl
import os
import glob
import base64

# Boeing color scheme
BOEING_BLUE = "#0033A0"
BOEING_LIGHT_BLUE = "#A0CCEB"
BOEING_GRAY = "#444444"

# Function to load and display the Boeing logo
def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def set_boeing_theme():
    # Note: You'll need to replace 'path_to_boeing_logo.png' with the actual path to a Boeing logo image file
    logo_path = 'Boeing_full_logo.png'
    logo_base64 = get_base64_of_bin_file(logo_path)
    
    st.markdown(
        f"""
        <style>
            .reportview-container .main .block-container{{
                max-width: 1000px;
                padding-top: 2rem;
                padding-bottom: 2rem;
                padding-left: 5rem;
                padding-right: 5rem;
            }}
            .reportview-container .main {{
                color: {BOEING_GRAY};
                background-color: white;
            }}
            .sidebar .sidebar-content {{
                background-color: {BOEING_LIGHT_BLUE};
            }}
            h1 {{
                color: {BOEING_BLUE};
            }}
            .stButton>button {{
                color: white;
                background-color: {BOEING_BLUE};
                border-radius: 5px;
            }}
            .stTextInput>div>div>input {{
                border-color: {BOEING_BLUE};
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )
    
    st.markdown(
        f"""
        <img src="data:image/png;base64,{logo_base64}" alt="Boeing Logo" style="width:200px;margin-bottom:20px;">
        """,
        unsafe_allow_html=True,
    )

# Set Boeing theme
set_boeing_theme()

st.title("RCR Sheet Automator")

# Create input fields
input1 = st.text_input("Enter the directory path where the files are stored (e.g., C:\\Users\\YourName\\Documents\\Files): ")
input2 = st.text_input("Enter the name of the permanent Excel file (e.g., permanent_file.xlsm): ")
input3 = st.text_input("Enter the pattern for the monthly files (e.g., groceries_*.xlsx): ")
input4 = st.text_input("Enter the name of the sheet in the permanent file where the data should be added: ")
input5 = st.text_input("Enter the starting row number (e.g., 4): ")

# Add a button to submit the inputs
if st.button("Submit"):
    if input1 and input2 and input3 and input4 and input5:
        try:
            # Set up variables
            directory_path = input1
            permanent_file_name = input2
            monthly_file_pattern = input3
            sheet_name = input4
            start_row = int(input5)
            header_row = 3

            # Create a progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()

            # Update progress
            def update_progress(progress, status):
                progress_bar.progress(progress)
                status_text.text(status)

            # Construct the full path for the permanent file
            permanent_file = os.path.join(directory_path, permanent_file_name)

            # Construct a file pattern to match the monthly files
            file_pattern = os.path.join(directory_path, monthly_file_pattern)

            update_progress(0.1, "Looking for monthly files...")

            # Find all files matching the pattern
            monthly_files = glob.glob(file_pattern)

            if not monthly_files:
                st.error("No monthly files found matching the pattern.")
                st.stop()

            # Find the latest file based on the naming convention
            latest_file = max(monthly_files, key=os.path.getctime)

            update_progress(0.2, f"Reading data from {latest_file}...")

            # Read the monthly data
            monthly_data = pd.read_excel(latest_file, engine='openpyxl')

            update_progress(0.3, "Opening permanent file...")

            # Open the permanent file and load the workbook
            permanent_wb = openpyxl.load_workbook(permanent_file, keep_vba=True, read_only=False, data_only=False)

            if sheet_name not in permanent_wb.sheetnames:
                st.error(f"Sheet '{sheet_name}' does not exist in the permanent file.")
                st.stop()

            sheet = permanent_wb[sheet_name]

            update_progress(0.4, "Processing headers...")

            # Get column headers
            permanent_headers = [str(cell.value) for cell in sheet[header_row] if cell.value is not None]
            monthly_headers = [header.strip().lower() for header in monthly_data.columns]

            update_progress(0.5, "Preparing data for update...")

            # Prepare data for update
            data_to_update = {}
            for header in permanent_headers:
                normalized_header = header.strip().lower()
                if normalized_header in monthly_headers:
                    monthly_col_idx = monthly_headers.index(normalized_header)
                    column_data = monthly_data.iloc[:, monthly_col_idx].tolist()
                    data_to_update[normalized_header] = column_data

            update_progress(0.6, "Clearing previous data...")

            # Clear previous data but keep formatting and formulas
            for col_idx, header in enumerate(permanent_headers):
                for row_idx in range(start_row, sheet.max_row + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx + 1)
                    if cell.data_type != 'f':  # Skip if cell contains a formula
                        cell.value = None

            update_progress(0.7, "Updating data...")

            # Update the sheet with new data
            for col_idx, header in enumerate(permanent_headers):
                normalized_header = header.strip().lower()
                if normalized_header in data_to_update:
                    column_data = data_to_update[normalized_header]
                    for row_idx, value in enumerate(column_data, start=start_row):
                        cell = sheet.cell(row=row_idx, column=col_idx + 1)
                        if cell.data_type != 'f':  # Skip if cell contains a formula
                            cell.value = value

            update_progress(0.9, "Saving permanent file...")

            # Save the updated permanent file
            permanent_wb.save(permanent_file)

            update_progress(1.0, "Process completed successfully!")
            st.success(f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{sheet_name}' starting at row {start_row}")

        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
    else:
        st.warning("Please fill in all fields before submitting.")
