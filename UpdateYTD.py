import pandas as pd
import openpyxl
import os
import glob

# Get input from the user
directory_path = input("Enter the directory path where the files are stored (e.g., C:\\Users\\YourName\\Documents\\Files): ")
permanent_file_name = input("Enter the name of the permanent Excel file (e.g., permanent_file.xlsm): ")
monthly_file_pattern = input("Enter the pattern for the monthly files (e.g., groceries_*.xlsx): ")
permanent_sheet_name = input("Enter the name of the sheet in the permanent file where the data should be added: ")
permanent_header_row = int(input("Enter the row number of the headers in the permanent file (e.g., 3): "))
permanent_start_row = int(input("Enter the starting row number to populate in the permanent file (e.g., 4): "))
temporary_header_row = int(input("Enter the row number of the headers in the temporary file (e.g., 2): "))
temporary_start_row = int(input("Enter the starting row number of the data in the temporary file (e.g., 3): "))

# Construct the full path for the permanent file
permanent_file = os.path.join(directory_path, permanent_file_name)

# Construct a file pattern to match the monthly files
file_pattern = os.path.join(directory_path, monthly_file_pattern)

# Print the file pattern to verify it's correct
print(f"Looking for files matching the pattern: {file_pattern}")

# Find all files matching the pattern
monthly_files = glob.glob(file_pattern)

# Print the list of files found to verify
print(f"Found monthly files: {monthly_files}")

# Ensure there are files matching the pattern
if not monthly_files:
    print("No monthly files found matching the pattern.")
    exit()

# Find the latest file based on the naming convention
latest_file = max(monthly_files, key=os.path.getctime)

# Verify that the latest file is a valid Excel file
try:
    monthly_data = pd.read_excel(latest_file, header=temporary_header_row - 1, engine='openpyxl')
except Exception as e:
    print(f"Error reading the monthly file: {e}")
    exit()

try:
    # Open the permanent file and load the workbook
    permanent_wb = openpyxl.load_workbook(permanent_file, keep_vba=True)
except Exception as e:
    print(f"Error opening the permanent file: {e}")
    exit()

# Access the specific sheet in the permanent workbook
if permanent_sheet_name not in permanent_wb.sheetnames:
    print(f"Sheet '{permanent_sheet_name}' does not exist in the permanent file.")
    exit()
sheet = permanent_wb[permanent_sheet_name]

# Get column headers from the permanent file and filter out non-string headers
permanent_headers = [cell.value for cell in sheet[permanent_header_row] if isinstance(cell.value, str)]
print(f"Permanent headers: {permanent_headers}")

# Get column headers from the monthly file and filter out non-string headers
monthly_headers = [header.strip().lower() for header in monthly_data.columns if isinstance(header, str)]
print(f"Monthly headers: {monthly_headers}")

# Clear the cells in the specified rows, but keep formatting and equations
for row in sheet.iter_rows(min_row=permanent_start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        if cell.data_type == 'f':  # If cell contains a formula, skip clearing
            continue
        cell.value = None

# Update the specific sheet starting at the specified row
for col_idx, header in enumerate(permanent_headers):
    normalized_header = header.strip().lower()
    if normalized_header in monthly_headers:
        monthly_col_idx = monthly_headers.index(normalized_header)
        print(f"Updating column: {header}")
        # Copy data starting from the specified temporary start row directly to the specified permanent start row
        temp_data = monthly_data.iloc[temporary_start_row - temporary_header_row:, monthly_col_idx]
        for row_offset, value in enumerate(temp_data, start=0):
            cell = sheet.cell(row=permanent_start_row + row_offset, column=col_idx + 1)
            cell.value = value

# Save the updated permanent file
try:
    permanent_wb.save(permanent_file)
except Exception as e:
    print(f"Error saving the permanent file: {e}")
    exit()

print(f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{permanent_sheet_name}' starting at row {permanent_start_row}'")
