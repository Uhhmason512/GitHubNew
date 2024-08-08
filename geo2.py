import pandas as pd
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import time
import openpyxl

# Load the Excel file using pandas
excel_file = pd.ExcelFile('RCR_Prioritization_File_06.26.2024_DM Review.xlsm')

# Read the Candidate_List tab
candidate_df = excel_file.parse('Top 30', header=2)

# Read the Facilitator List tab
facilitator_df = excel_file.parse('Facilitator List With Location')

# Create a new column in the Candidate_List tab to store the closest facilitator name
candidate_df['Closest Facilitator'] = ''
geolocator = Nominatim(user_agent='my_app')

def geocode_with_retry(location):
    max_retries = 5
    retries = 0
    while retries < max_retries:
        try:
            return geolocator.geocode(location)
        except GeocoderTimedOut:
            retries += 1
            time.sleep(1)  # Wait for 1 second before retrying
    return None

facilitator_locations = {}
facilitator_count = {}
facilitator_df = facilitator_df.dropna()
for facilitator_index, facilitator_row in facilitator_df.iterrows():
    facilitator_location = facilitator_row['Location']
    facilitator_name = facilitator_row['Name']
    print(facilitator_location)
    facilitator_state, facilitator_city = facilitator_location.split('-')
    facilitator_location = f"{facilitator_city.strip()}, {facilitator_state.strip()}"
    facilitator_coordinates = geocode_with_retry(facilitator_location)
    facilitator_locations[facilitator_name] = facilitator_coordinates
    facilitator_count[facilitator_name] = 0

# Iterate over each RCR in the Candidate_List tab
for index, candidate_row in candidate_df.head(30).iterrows():
    candidate_location = f"{candidate_row['City']}, {candidate_row['State']}"
    candidate_coordinates = geocode_with_retry(candidate_location)
    candidate_latitude = candidate_coordinates.latitude
    candidate_longitude = candidate_coordinates.longitude
    closest_facilitator = None
    min_distance = float('inf')
    
    # Iterate over each facilitator in the Facilitator List tab
    for facilitator_index, facilitator_row in facilitator_df.iterrows():
        facilitator_coordinates = facilitator_locations[facilitator_row['Name']]
        facilitator_latitude = facilitator_coordinates.latitude
        facilitator_longitude = facilitator_coordinates.longitude
        
        distance = geodesic((candidate_latitude, candidate_longitude), (facilitator_latitude, facilitator_longitude)).miles
        
        # Update the closest facilitator if a closer one is found
        if distance < min_distance and facilitator_count[facilitator_row['Name']] < facilitator_row['RCR_Limit']:
            min_distance = distance
            closest_facilitator = facilitator_row['Name']
    
    # Assign the closest facilitator name to the corresponding row in the Candidate_List tab
    candidate_df.at[index, 'Closest Facilitator'] = closest_facilitator
    facilitator_count[closest_facilitator] = facilitator_count[closest_facilitator] + 1

# Load the original workbook with openpyxl
wb = openpyxl.load_workbook('RCR_Prioritization_File_06.26.2024_DM Review.xlsm', keep_vba=True)

# Select the sheet you want to update
ws = wb['Top 30']

# Add the new column header
ws.cell(row=3, column=ws.max_column + 1, value='Closest Facilitator')

# Write the data to the new column
for idx, value in enumerate(candidate_df['Closest Facilitator'], start=4):
    ws.cell(row=idx, column=ws.max_column, value=value)

# Save the workbook
wb.save('RCR_Prioritization_File_06.26.2024_DM Review.xlsm')