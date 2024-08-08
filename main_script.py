import os
import sys
import time
import json
import pandas as pd
import openpyxl
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import multiprocessing

def start_streamlit():
    command = f"streamlit run streamlit_app.py"
    process = subprocess.Popen(command, shell=True)
    process.communicate()

def run_main_logic():
    try:
        while not os.path.exists('trigger.txt'):
            time.sleep(1)

        with open('params.json', 'r') as f:
            params = json.load(f)

        directory_path = params["directory_path"]
        permanent_file_name = params["permanent_file_name"]
        monthly_file_pattern = params["monthly_file_pattern"]
        sheet_name = params["sheet_name"]
        start_row = int(params["start_row"])
        header_row = 3

        os.remove('trigger.txt')

        permanent_file = os.path.join(directory_path, permanent_file_name)
        file_pattern = os.path.join(directory_path, monthly_file_pattern)

        monthly_files = glob.glob(file_pattern)

        if not monthly_files:
            print("No monthly files found matching the pattern.")
            return

        latest_file = max(monthly_files, key=os.path.getctime)
        monthly_data = pd.read_excel(latest_file, engine='openpyxl')
        permanent_wb = openpyxl.load_workbook(permanent_file, keep_vba=True, read_only=False, data_only=False)

        if sheet_name not in permanent_wb.sheetnames:
            print(f"Sheet '{sheet_name}' does not exist in the permanent file.")
            return

        sheet = permanent_wb[sheet_name]
        permanent_headers = [cell.value for cell in sheet[header_row] if cell.value is not None]
        monthly_headers = [header.strip().lower() for header in monthly_data.columns]

        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.data_type != 'f':
                    cell.value = None

        for col_idx, header in enumerate(permanent_headers):
            normalized_header = header.strip().lower()
            if normalized_header in monthly_headers:
                monthly_col_idx = monthly_headers.index(normalized_header)
                for row_idx, value in enumerate(monthly_data.iloc[:, monthly_col_idx]):
                    cell = sheet.cell(row=row_idx + start_row, column=col_idx + 1)
                    cell.value = value
                if normalized_header == "on rcr events report":
                    break

        print(f"Permanent file '{permanent_file_name}' updated with data from '{latest_file}' in sheet '{sheet_name}' starting at row {start_row}")

        excel_file = pd.ExcelFile(permanent_file)
        candidate_df = excel_file.parse('Top 30', header=2)
        facilitator_df = excel_file.parse('Facilitator List With Location')
        candidate_df['Closest Facilitator'] = ''
        geolocator = Nominatim(user_agent='my_app')

        def geocode_with_retry(location):
            max_retries = 5
            retries = 0
            while retries < max_retries:
                try:
                    return geolocator.geocode(location)
                except GeocoderTimedOut:
                    retries += 1
                    time.sleep(1)
            return None

        facilitator_locations = {}
        facilitator_count = {}
        facilitator_df = facilitator_df.dropna()
        for facilitator_index, facilitator_row in facilitator_df.iterrows():
            facilitator_location = facilitator_row['Location']
            facilitator_name = facilitator_row['Name']
            facilitator_state, facilitator_city = facilitator_location.split('-')
            facilitator_location = f"{facilitator_city.strip()}, {facilitator_state.strip()}"
            facilitator_coordinates = geocode_with_retry(facilitator_location)
            facilitator_locations[facilitator_name] = facilitator_coordinates
            facilitator_count[facilitator_name] = 0

        for index, candidate_row in candidate_df.head(30).iterrows():
            candidate_location = f"{candidate_row['City']}, {candidate_row['State']}"
            candidate_coordinates = geocode_with_retry(candidate_location)
            if candidate_coordinates is None:
                continue
            candidate_latitude = candidate_coordinates.latitude
            candidate_longitude = candidate_coordinates.longitude
            closest_facilitator = None
            min_distance = float('inf')

            for facilitator_index, facilitator_row in facilitator_df.iterrows():
                facilitator_coordinates = facilitator_locations[facilitator_row['Name']]
                if facilitator_coordinates is None:
                    continue
                facilitator_latitude = facilitator_coordinates.latitude
                facilitator_longitude = facilitator_coordinates.longitude

                distance = geodesic((candidate_latitude, candidate_longitude), (facilitator_latitude, facilitator_longitude)).miles

                if distance < min_distance and facilitator_count[facilitator_row['Name']] < facilitator_row['RCR_Limit']:
                    min_distance = distance
                    closest_facilitator = facilitator_row['Name']

            if closest_facilitator is not None:
                candidate_df.at[index, 'Closest Facilitator'] = closest_facilitator
                facilitator_count[closest_facilitator] = facilitator_count.get(closest_facilitator, 0) + 1
            else:
                print(f"No facilitator found for candidate at index {index}")

        ws = permanent_wb['Top 30']
        ws.cell(row=3, column=ws.max_column + 1, value='Closest Facilitator')

        for idx, value in enumerate(candidate_df['Closest Facilitator'], start=4):
            ws.cell(row=idx, column=ws.max_column, value=value)

        permanent_wb.save(permanent_file)
        print("Process completed successfully.")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

def main():
    if multiprocessing.current_process().name == "MainProcess":
        print("Starting Streamlit app...")
        p = multiprocessing.Process(target=start_streamlit)
        p.start()
        run_main_logic()
        p.join()

if __name__ == "__main__":
    main()
